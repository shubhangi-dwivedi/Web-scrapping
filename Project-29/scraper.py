import json
import requests
import xlsxwriter
from bs4 import BeautifulSoup
from flask_restful import Resource
from urllib.request import urlopen
import pandas as pd
import openpyxl
from textblob import TextBlob
import re
import openpyxl.utils.exceptions


#1
class anOtherScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")
                URL="https://www.anothermag.com/search?q="+seed
                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1= soup.find_all("div",class_="article-list-item list-item")

                    data = []
                    for d in data1:
                        short_description = d.find("a", class_="section-link")
                        title = d.find("a", class_="title")
                        span_text=d.find("span",class_="nowrap")
                        url = "https://www.anothermag.com" + title["href"]
                        #snippet = None

                        short_description = short_description.text.strip("\n")

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    url = json_without_slash
                    s1 = json.dumps(json_without_slash)
                    # d2 = json.loads(s1)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)
                    wb.save('anOtherScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True
        #---------------------------

#2
class architecturalDigestScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")
                URL = "https://www.architecturaldigest.in/search/?q="+ seed.upper() +"&sort=score+desc"
                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemContent-gYsIAS eobScx summary-item__content")  #-------------

                    data = []
                    for d in data1:
                        short_description = d.find("span", class_="RubricName-eXGqmo dVwJRo")
                        title = d.find("h2")
                        snippet = d.find("div", class_="BaseWrap-sc-UrHlS BaseText-fFrHpW SummaryItemDek-dwcsSh boMZdO dHFxXE eWmA-dv summary-item__dek")
                        url = "https://www.architecturaldigest.in" + d.find("a", class_= "summary-item__hed-link")["href"]

                        short_description = short_description.text.strip("\n")

                        data.append({"short_description": short_description, "title": title.text, "snippet":snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('architecturalDigestScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True
            #------------

#3
class countryLivingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")
                URL ="https://www.countryliving.com/search/?q=" + seed
                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")  # -------------

                    data = []
                    for d in data1:
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.countryliving.com" + short_url
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")

                        '''
                        if short_url is None:
                            continue
                        if title.text is None:
                            continue
                        if snippet.text is None:
                            continue
                            '''
                        snippet=snippet.text.strip("\n")
                        short_description = short_description.strip("\n")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('countryLivingScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True
        # ------------

#4
class elleDecorScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")
                URL = "https://www.elledecor.com/search/?q=" + seed
                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")  # -------------

                    data = []
                    for d in data1:
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.elledecor.com" + short_url
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")

                        snippet = snippet.text.strip("\n")
                        short_description = short_description.strip("\n")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('elleDecorScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

#5
class femina(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed=seed.strip("\n")
                seed = seed.strip()
                seed=seed.replace(" ","+")

                URL="https://www.femina.in/search/tag_"+ seed +"&sort=score+desc?pg=10"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1= soup.find_all("div", class_="search-section")

                    data=[]
                    for d in data1:
                        short_description = d.find("div", class_="clearfix")
                        title = d.find("h3")
                        snippet = d.find("p")
                        url = title.find("a")["href"]

                        short_description=short_description.text.strip("\n")

                        data.append({"short_description":short_description,  "title":title.text ,"snippet": snippet.text, "url":url})

                        data2={"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('femina.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True


#6
class foodAndWineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.foodandwine.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        url=d['href']
                        cat_tag = d.find("div", class_="card__content")["data-tag"]
                        title = d.find("span", class_="card__title-text")

                        data.append({"cat_tag": cat_tag, "title": title.text, "url":url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['cat_tag'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('foodAndWineScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#7
class gfmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.gfmag.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="article-preview search-result")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.gfmag.com" + short_url
                        snippet = d.find_all("p")[1]
                        snippet=snippet.text.strip("...")
                        snippet=snippet.replace("\r\n\r\n"," ")
                        title=title.text.strip("\n")

                        short_description = short_description.strip("\n")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gfmagScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True


#8
class hotelierScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.hoteliermagazine.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title")
                        snippet = d.find("div", class_="td-excerpt").text
                        url = title.find("a")["href"]


                        snippet=snippet.lstrip("\r\n                    \n")
                        snippet=snippet.rstrip("...                ")
                        snippet=snippet.replace("\n","")

                        data.append({"title": title.text, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('hotelierScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#9
class houseBeautifulScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.housebeautiful.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", "simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        snippet =snippet.text.replace("\n", " ")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.housebeautiful.com" + short_url

                        data.append({"short_description":short_description, "title": title.text, "snippet": snippet, "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('houseBeautifulScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True


#10
class lodgingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://lodgingmagazine.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td-block-span6")

                    data = []
                    for d in data1:
                        short_description = d.find("a", class_="td-post-category")
                        title = d.find("h3")
                        snippet = d.find("div", class_="td-module-meta-info")
                        snippet=snippet.text.strip("\n")
                        url = title.find("a")["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('lodgingScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#11
class nationalGeographicScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.nationalgeographic.com/search?q=" + seed + "&location=srp&short_description=manual"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="ResultCard")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="ResultCard__Title")
                        description = d.find("span", class_="ResultCard__Description")
                        snippet = description.find("span", class_="RichText")
                        url = d.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('nationalGeographicScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#12
class nylonScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.nylon.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="ofI")

                    data = []
                    for d in data1:
                        short_description = d.find("p", class_="pxx")
                        title = d.find("p", class_="icJ")
                        url = "https://www.nylon.com" + d["href"]

                        data.append({"short_description": short_description.text, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('nylonScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#13
class romanticHomesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.romantichomes.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="post-list-bdr cf")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="excerpt ex-wd").find("a")
                        url = title["href"]
                        short_description = url.split("/")[3].replace("-", " ") + " " + url.split("/")[4].replace("-", " ")

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('romanticHomesScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#14
class theSpruceScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.thespruce.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="card-list__entry")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="card__content").get('data-tag')
                        title = d.find("span", class_="card__title")
                        url = d.find("a", class_="card-list__card")["href"]

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theSpruceScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#15
class verandaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.veranda.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title item-title")
                        snippet =d.find("div",class_="simple-item-dek item-dek")
                        short_url =d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.veranda.com" + short_url

                        snippet=snippet.text.strip()

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('verandaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#16
class instoreScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://instoremag.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="mvp-blog-story-out")

                    data = []
                    for d in data1:
                        link = d.find_all("a")[2]
                        short_description = d.find("div", class_="mvp-cat-date-wrap").find("a")
                        title = d.find("h2")
                        url = link["href"]
                        snippet = d.find("p")
                        snippet=snippet.text.strip("...")
                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('instoreScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#17
class gjepcScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://gjepc.org/solitaire/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find("div", class_="small-12 medium-8 columns").find_all("article", class_= "post")

                    data = []

                    if len(data1)!=0:
                        for d in data1:
                            short_description = d.find("aside", class_="post-meta")
                            title = d.find("div", class_="post-title")
                            url = title.find("a")["href"]
                            snippet = d.find("p")

                            title = title.text.replace("\n", "")
                            title = title.replace("\t", "")
                            title = title.strip()
                            snippet = snippet.text.replace("\n", "")
                            snippet = snippet.replace("\t", "")
                            snippet = snippet.rstrip("…")
                            short_description=short_description.text.replace("\n", "")
                            short_description=short_description.replace("\t", "")
                            short_description = short_description.strip()

                            data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                         "url": url})

                            data2 = {"keyword": seed, "data": data}
                        jsonData = json.dumps(data2)
                        json_without_slash = json.loads(jsonData)

                        s1 = json.dumps(json_without_slash)

                        json_object = json.loads(s1)
                        for each in json_object['data']:
                            l = []
                            l.append(json_object['keyword'])
                            l.append(each['title'])
                            l.append(each['short_description'])
                            l.append(each['snippet'])
                            l.append(each['url'])

                            x = each['title'].replace("...", "")
                            title_noun = TextBlob(x)
                            title_noun = title_noun.noun_phrases
                            l.append(str(title_noun))

                            data = tuple(l)
                            ws.append(data)

                        wb.save('gjepcScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#18
class flowerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://flowermag.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find("div", class_="blog blog-style-3x blog-style-masonry").find_all("div", class_= "description")

                    data = []
                    if len(data1) != 0:
                        for d in data1:
                            title = d.find("div", class_="post-title")
                            url = title.find("a")["href"]
                            snippet = d.find("div", class_="post-text")

                            title=title.text.replace("\n","")
                            title=title.replace("\t","")
                            title=title.strip()
                            snippet=snippet.text.replace("\n","")
                            snippet=snippet.replace("\t","")
                            snippet=snippet.strip()

                            data.append({"title": title, "snippet": snippet,
                                         "url": url})

                            data2 = {"keyword": seed, "data": data}
                        jsonData = json.dumps(data2)
                        json_without_slash = json.loads(jsonData)

                        s1 = json.dumps(json_without_slash)

                        json_object = json.loads(s1)
                        for each in json_object['data']:
                            l = []
                            l.append(json_object['keyword'])
                            l.append(each['title'])
                            l.append(each['snippet'])
                            l.append(each['url'])

                            x = each['title'].replace("...", "")
                            title_noun = TextBlob(x)
                            title_noun = title_noun.noun_phrases
                            l.append(str(title_noun))

                            data = tuple(l)
                            ws.append(data)

                        wb.save('flowerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#19
class elleDecor2Scraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://elledecor.in/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_= "coloum col-sm-4")

                    data = []
                    for d in data1:
                        title = d.find("h4")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        data.append({"short_description": short_description,"title": title.text, "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('elleDecor2Scraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#20
class allureScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.allure.com/search?q=" + seed + "&sort=score+desc"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_= "summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h3")
                        short_url=d.find("a", class_="summary-item__hed-link")["href"]
                        url= "https://www.allure.com"+short_url
                        snippet = d.find("div", class_="summary-item__dek")

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text, "url":url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('allureScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#21
class stampingtonScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://stampington.com/search.php?search_query_adv=" + seed + "#search-results-information"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="card card-gallery card-hover")

                    data = []
                    for d in data1:
                        title = d.find("a", class_="card-ellipsis").span
                        url = d.find("a", class_="card-ellipsis")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('stampingtonScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#22
class indianJewelerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://indianjeweller.in/search-news/" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", "news-row")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="title-16")
                        snippet = d.find("div", class_="desc-text")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="post__small-text-meta")

                        title=title.text.replace("\n","")
                        snippet = snippet.text.replace("\n", "")
                        short_description = short_description.text.replace("\n", "")
                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('indianJewelerScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#23
class metropolisScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://metropolismag.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="card__metro_project")

                    data = []
                    for d in data1:
                        paras = d.find_all("p")
                        title = d.find("a", class_="post-cards__title-link")
                        snippet = paras[1]
                        url = title["href"]
                        short_description = paras[0]

                        title = title.text.replace("\n", "")
                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.strip()
                        snippet = snippet.lstrip("… ")
                        snippet = snippet.rstrip("… ")

                        data.append({"short_description": short_description.text, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('metropolisScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#24
class moneyScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://money.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="search-result")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="headline")
                        url = "https://money.com" + title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        data.append({"short_description":short_description,"title": title.text, "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('moneyScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#25
class motoringScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://motoringworld.in/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="card")

                    data = []
                    for d in data1:
                        content =d.find("div", class_="content")
                        title = content.find("div", class_="title")
                        url = title.find("a")["href"]
                        short_description = content.find("div", class_="section")

                        short_description=short_description.text.lstrip("\n")
                        short_description = short_description.rstrip(" \n")

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('motoringScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#26
class professionalWomanScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://professionalwomanmag.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="type-post")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="entry-title")
                        snippet = d.find("div", class_="entry-content").find_all("p")[-1]
                        url = snippet.find("a", class_="more-link")["href"]
                        short_description = d.find("div", class_="entry-meta")

                        short_description=short_description.text.strip('\n')
                        snippet=snippet.text.rstrip(" →")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('professionalWomanScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#27
class townAndCountryScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.townandcountrymag.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        short_url = d.find("a")["href"]
                        url = "https://www.townandcountrymag.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ") + short_url.split("/")[2].replace("-",
                                                                                                                        " ")
                        snippet=snippet.text.strip("\n")
                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('townAndCountryScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#28
class travelAndLeisureIndiaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.travelandleisureindia.in/?s=" + seed + "&type=&category=&sort="

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="card")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="categories")
                        url = d.find("a")["href"]
                        snippet = url.split("/")[5].replace("-", " ")
                        short_description = url.split("/")[3].replace("-", " ") + url.split("/")[4].replace("-", " ")

                        title=title.text.lstrip("\n\n\t\t\t\t")
                        title=title.rstrip("\n\t\t\t\n")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('travelAndLeisureIndiaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#29
class travelPeacockScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.travelpeacockmagazine.com/?s="+ seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="post-title")
                        url = d.find("a")["href"]
                        snippet = d.find("div", class_="post-content")
                        short_description = d.find("div", class_="thb-post-top")

                        short_description=short_description.text.replace("\n","")
                        short_description=short_description.replace("\t","")
                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet=snippet.replace("…","")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('travelPeacockScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#30
class vanityFairScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.vanityfair.com/search?q="+ seed + "&sort=score+desc"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo hsRhGO summary-item summary-item--has-border summary-item--article summary-item--no-icon summary-item--text-align-left summary-item--layout-placement-side-by-side-desktop-only summary-item--layout-position-image-left summary-item--layout-proportions-50-50 summary-item--side-by-side-align-center summary-item--side-by-side-image-right-mobile-false summary-item--standard SummaryItemWrapper-bGtGFH klkoMz summary-list__item")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h2")
                        snippet = d.find("div", class_="summary-item__dek")

                        url = "https://www.vanityfair.com" + d.find("a", class_="summary-item__hed-link")["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('vanityFairScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#31
class womensHealthScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.womenshealthmag.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        snippet = snippet.text.strip("\n")

                        short_url = d.find("a")["href"]
                        url = "https://www.womenshealthmag.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('womensHealthScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#32
class theJewelryScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.thejewelrymagazine.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="td-module-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="td-module-meta-info")

                        short_description=short_description.text.strip("\n")
                        snippet=snippet.text.replace("\r\n"," ")
                        snippet=snippet.strip()
                        snippet=snippet.rstrip("...")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theJewelryScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#33
class theEnglishGardenScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.theenglishgarden.co.uk/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="td-module-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        snippet=snippet.text.strip("\n")
                        snippet=snippet.strip()
                        snippet = snippet.strip("...")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theEnglishGardenScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#34
class teenVogueScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.teenvogue.com/search?q=" + seed + "&sort=score+desc"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("a", class_="RubricLink-DDpgX kWyAIu rubric__link")
                        title = d.find("h2", class_="summary-item__hed")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.teenvogue.com" + d.find("a", class_="summary-item__hed-link")["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                         "url": url})

                        data2 = {"keyword": seed, "data": data}
                        #print(data2)

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)


                    for each in json_without_slash['data']:
                        l = []
                        l.append(json_without_slash['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)

                        ws.append(data)

                    wb.save('teenVogueScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass

        wb.close()
        return True

#35
class solitaireScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.solitairemagazine.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="entry-item")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="entry-meta")
                        title = d.find("h2", class_="entry-title")
                        snippet = d.find("section", class_="entry-content")
                        url = title.find("a")["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('solitaireScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#36
class southernLivingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.southernliving.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    #print(data1)
                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", "/")

                        print(title.text)
                        print(short_description)
                        print(url)

                        title = title.text.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.strip()

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                        print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('southernLivingScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#37
class styleAtHomeScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.styleathome.com/search?search[query]=" + seed + "&search[sort]="

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="result")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="title")
                        short_url = title.find("a")["href"]
                        url = "https://www.styleathome.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ") + " " + short_url.split("/")[
                            2].replace("-", " ")

                        title=title.text.strip("\n")

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('styleAtHomeScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#38
class realSimpleScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.realsimple.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", "/")

                        title = title.text.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.strip()

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('realSimpleScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#39
class realHomesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.realhomes.com/search?searchTerm=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="listing__item")

                    data = []
                    for d in data1:
                        url = d.find("a", class_="listing__link")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        title = d.find("h2", class_="listing__title")
                        snippet = d.find("p", class_="listing__text--strapline")

                        title = title.text.strip("\n")
                        snippet = snippet.text.strip("\n")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('realHomesScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#40
class newYorkerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.newyorker.com/search/q/" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="River__riverItem___3huWr")

                    data = []
                    for d in data1:
                        links = d.find_all("a")
                        short_description = links[0]
                        title = d.find("h4", class_="River__hed___re6RP")
                        snippet = d.find("h5", class_="River__dek___CayIg")
                        url = "https://www.newyorker.com" + links[-2]["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('newYorkerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

#41
class moneySenseScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:
                print(seed)

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.moneysense.ca/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="post--item")

                    data = []
                    for d in data1:
                        #print(seed)

                        short_description = d.find("p", class_="post-categories uppercase").text.strip()
                        title = d.find("h3").text
                        snippet = d.find("p", class_="excerpt")
                        url = d.find("a", class_="post--link")["href"]

                        #if snippet!="":
                         #   snippet=snippet.text.strip("...")
                        #print(snippet)
                        data.append({"short_description": short_description, "title": title, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('moneySenseScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#42
class maximScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.maxim.com/?s=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card-title")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        title=title.text.strip("\n")

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('maximScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#43
class livingetcScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.livingetc.com/search?searchTerm=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="listing__item")

                    data = []
                    for d in data1:
                        url = d.find("a", class_="listing__link")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        title = d.find("h2", class_="listing__title")
                        snippet = d.find("p", class_="listing__text")

                        title=title.text.strip("\n")
                        snippet=snippet.text.strip("\n")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('livingetcScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#44
class inStyleScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.instyle.com/search?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        title = title.text.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.strip()

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('inStyleScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#45
class harpersBazaarScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.harpersbazaar.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ") + " " + short_url.split("/")[
                            2].replace("-", " ")
                        url = "https://www.harpersbazaar.com" + short_url

                        snippet=snippet.text.strip("\n")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('harpersBazaarScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#46
class graziaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.grazia.co.in/search/" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="thumbnail")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        url = title.find("a")["href"]
                        snippet = url.split("/")[3].replace("-", " ")

                        data.append({"title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('graziaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#47
class glamourScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.glamourmagazine.co.uk/search?q=" + seed + "&sort=score+desc"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h3")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.glamourmagazine.co.uk" + d.find("a", class_="summary-item__hed-link")[
                            "href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('glamourScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#48
class gardensIllustratedScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.gardensillustrated.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="template-search-universal__card")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="standard-card-new__display-title")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.gardensillustrated.com" + short_url

                        title=title.text.replace("\n","")
                        title=title.strip()
                        title = re.sub(' +', ' ', title)

                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gardensIllustratedScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#49
class dwellScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.dwell.com/query/" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="_3mbrXRthE9F9t6iCkqG6F-")

                    data = []
                    for d in data1:
                        title = d
                        snippet = d.parent.find("div", class_="_37azNHMqzIJRxpdnCtt8J5")
                        url = "https://www.dwell.com" + title["href"]

                        data.append({"title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('dwellScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#50
class cosmopolitanScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.cosmopolitan.com/search/?q=" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.cosmopolitan.com" + short_url

                        snippet=snippet.text.strip("\n")

                        data.append({"short_description": short_description, "title": title.text, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cosmopolitanScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#51
class cntravellerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.cntraveller.in/search/?q=" + seed + "&sort=score+desc"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h2")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.cntraveller.in" + d.find("a", class_="summary-item__hed-link")["href"]

                        data.append({"short_description": short_description.text, "title": title.text, "snippet": snippet.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cntravellerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#52
class businessTodayScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.businesstoday.in/topic/" + seed

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find("ul", id="more_content_container").find_all("a")

                    data = []
                    for d in data1:
                        title = d.find("h2")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ") + " " + url.split("/")[4].replace("-", " ")

                        data.append({"short_description": short_description, "title": title.text,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('businessTodayScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

#53
class bhgScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.strip("\n")
                seed = seed.strip()
                seed = seed.replace(" ", "+")

                URL = "https://www.bhg.com/search/?q=" + seed + "&quotequery=&mod=DNH_S"

                page = requests.get(URL)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        title = title.text.replace("\n", "")
                        title=title.replace("\t","")
                        title = title.strip()



                        data.append({"short_description": short_description, "title": title,
                                     "url": url})

                        data2 = {"keyword": seed, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('bhgScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True