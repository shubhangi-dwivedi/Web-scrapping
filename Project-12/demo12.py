import string

import wordninja
from bs4 import BeautifulSoup
import requests as rq
from openpyxl.reader.excel import load_workbook
import operator
import xlsxwriter

import openpyxl

file1 = load_workbook(filename='python_job_records.xlsx')
ws=file1['skills_records']

start_col = 1
end_col = 1

pg=0

workbook= xlsxwriter.Workbook("all_records"+".xlsx")
worksheet= workbook.add_worksheet("all_jobs")

worksheet.write(0,0,'#')
worksheet.write(0,1,'Company_name')
worksheet.write(0,2,'skills')
worksheet.write(0,3,'description')

for i in range(2, ws.max_row+1):
    row = [cell.value for cell in ws[i][start_col:end_col+1]]
    print(row)

    string1=' '.join(row)
    string1=string1.replace(" ", "+")
    print(string1)

    skill=string1
    print(skill)


    for page in range(1, 6):
        pg_no = str(page)
        url = 'https://www.timesjobs.com/candidate/job-search.html?from=submit&actualTxtKeywords=' + skill + '&searchBy=0&rdoOperator=OR&searchType=personalizedSearch&luceneResultSize=25&postWeek=60&txtKeywords=' + skill + '&pDate=I&sequence=' + pg_no + '&startPage=1'
        #print(url)

        #print(page)
        html_text = rq.get(url + str(page) + '&startPage=1').text

        soup = BeautifulSoup(html_text, 'lxml')
        jobs = soup.find_all('li', class_='clearfix job-bx wht-shd-bx')

        #print(pg)
        if page > 1:
            pg += 25

        for index, job in enumerate(jobs):
            #print(pg)

            if job.find('h3', class_='joblist-comp-name'):
                company_name = job.find('h3', class_='joblist-comp-name').text.replace(' ', '')

            if job.find('h2', class_='joblist-comp-name'):
                company_name = job.find('h2', class_='joblist-comp-name').text.replace(' ', '')

            skills = job.find('span', class_='srp-skills').text.replace(' ', '')
            job_descrp = job.find('ul', class_='list-job-dtl clearfix').text.replace(' ', '')
            #print(job_descrp)
            #----------------------------

            sub1 = "JobDescription:"
            sub2 = '...'
            idx1 =job_descrp.index(sub1)
            idx2 = job_descrp.index(sub2)

            res = ''
            # getting elements in between
            for idx in range(idx1 + len(sub1) + 1, idx2):
                res = res + job_descrp[idx]

            # print(res)

            ans = wordninja.split(res)
            # print(ans)

            final_string = ""
            for i in range(len(ans)):
                final_string = final_string + ans[i] + " "

            #------------------------------

            worksheet.write(pg + index + 1, 0, str(index + 1 + pg))
            worksheet.write(pg + index + 1, 1, company_name.strip())
            worksheet.write(pg + index + 1, 2, skills.strip())
            worksheet.write(pg + index + 1, 3, final_string)

workbook.close()