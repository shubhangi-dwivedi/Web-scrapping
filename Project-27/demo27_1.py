import re
from openpyxl.reader.excel import load_workbook
import nltk
import spacy
import xlsxwriter
from textblob import TextBlob
from nltk.stem import WordNetLemmatizer


nltk.download('wordnet')
nltk.download('omw-1.4')



#---------
def get_wordnet_pos(word):
    """Map POS tag to first character lemmatize() accepts"""
    tag = nltk.pos_tag([word])[0][1][0].upper()
    tag_dict = {"J": wordnet.ADJ,
                "N": wordnet.NOUN,
                "V": wordnet.VERB,
                "R": wordnet.ADV}

    return tag_dict.get(tag, wordnet.NOUN)
#------

nlp = spacy.load("en_core_web_sm")
nltk.download('averaged_perceptron_tagger')
from nltk.corpus import wordnet


file1 = load_workbook(filename='related_words1.xlsx')
ws=file1['synon']


start_col1 = 1
end_col1 = 4

workbook = xlsxwriter.Workbook("lemma_pos_related_words1" + ".xlsx")
worksheet = workbook.add_worksheet("lemma_pos")
worksheet.write(0, 0, '#')
worksheet.write(0, 1, 'word')
worksheet.write(0, 2, 'lemma')
worksheet.write(0, 3, 'pos')
worksheet.write(0, 4, 'ner')


index=0

for i in range(2,ws.max_row+1):
    row = [cell.value for cell in ws[i][start_col1:end_col1+1]]
    print(row)
    print(type(row))

    for temp in row:
        l=[]
        temp= temp.rstrip(']')
        temp= temp.lstrip('[')
        temp = temp.replace("'", "")

        l = list(temp.split(','))

        print(l)

        for x in l:
            if len(x.split())==1:

                x=x.strip()

                #pos
                blob_object = TextBlob(x)
                pos_res = blob_object.tags

                #lemm
                # 1. Init Lemmatizer
                lemmatizer = WordNetLemmatizer()

                lemma=lemmatizer.lemmatize(x, get_wordnet_pos(x))

                #ner
                l2 = []

                res = re.sub(r'[.,"\?:!;]', '', x)
                doc = nlp(res)
                entities = []
                labels = []

                print(doc.ents)

                for v in doc.ents:
                    entities.append(v)
                    labels.append(v.label_)

                l2 = list(zip(entities, labels))

                index += 1
                print("index is : ", index)
                print(pos_res,",",lemma,",",l2)

                worksheet.write(index, 0, index)
                worksheet.write(index, 1, x)
                worksheet.write(index, 2, lemma)
                worksheet.write(index, 3, str(pos_res))
                worksheet.write(index, 4, str(l2))


workbook.close()
