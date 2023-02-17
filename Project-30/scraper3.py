import json
import requests
import xlsxwriter
from bs4 import BeautifulSoup
from flask_restful import Resource
from urllib.request import urlopen
import pandas as pd
import openpyxl
from textblob import TextBlob
import re
import openpyxl.utils.exceptions
import wordninja
import unicodedata

class architectsJournalScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.architectsjournal.co.uk/?orderby=relevance&s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="item-list")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="post-title")
                        url = d.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet= d.find("div", class_="entry")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('architectsJournalScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class biospectrumasiaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "%20")

                URL = "https://biospectrumasia.com/search/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1_temp = soup.find("ul", class_="listingli nopadding")
                    data1 = data1_temp.find_all("li")

                    data = []
                    for d in data1:
                        title = d.find("p", class_="news-heading listing nomargine")
                        short_url = title.find("a")["href"]
                        short_description = d.find("p", class_="listingtitle nomargine")
                        url = "https://biospectrumasia.com"+short_url
                        snippet = d.find("p", class_="listingtext")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        x = short_description.split("|")
                        x=x[0]


                        data.append({"title": title,"short_description": x, "snippet": snippet ,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('biospectrumasiaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class biospectrumindiaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "%20")

                URL = "https://biospectrumindia.com/search/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1_temp = soup.find("ul", class_="listingli nopadding")
                    data1 = data1_temp.find_all("li")

                    data = []
                    for d in data1:
                        title = d.find("p", class_="news-heading listing nomargine")
                        short_url = title.find("a")["href"]
                        short_description = d.find("p", class_="listingtitle nomargine")
                        url = "https://biospectrumasia.com"+short_url
                        snippet = d.find("p", class_="listingtext")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        x = short_description.split("|")
                        x=x[0]


                        data.append({"title": title,"short_description": x, "snippet": snippet ,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('biospectrumindiaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class restechScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://restechtoday.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="row bottom-margin")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        snippet = d.find("span", class_="text")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="thumb-wrap relative")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('restechScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class oceanhomeScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.oceanhomemag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="cb-blog-style-a")

                    data = []
                    for d in data1:
                        title = d.find("h2", "cb-post-title")
                        snippet = d.find("div", class_="cb-excerpt")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet = snippet.replace("...  Read More >", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('oceanhomeScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class gardengateScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.gardengatemagazine.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="row my-3 py-3")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        snippet = d.find("summary")
                        url = "https://www.gardengatemagazine.com" + title.find("a")["href"]
                        short_description = d.find("h6", class_="pre")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.text.replace("| ", "")
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gardengateScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class renovatemagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://renovate-mag.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="result first item-article")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="title")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://renovate-mag.com/" + short_url
                        snippet = d.find("div", class_="rte clearfix")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('renovatemagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class barandkitchenmagazineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://barandkitchenmagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="overflow-hidden")

                    data = []
                    for d in data1:
                        short_description = d.find("div",
                                                   class_="text-white relative uppercase bg-black inline-block text-sm mx-auto -mt-3 -mb-3 px-4 py-1 z-20")
                        title = d.find("h2", class_="entry-title font-serif font-bold text-2xl mb-4")
                        url = title.find("a")["href"]
                        snippet = d.find("p", class_="mb-4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('barandkitchenmagazineScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class granddesignsmagazineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.granddesignsmagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        short_description = d.find("p", class_="the_category")
                        title = d.find("h2")
                        url = title.find("a")["href"]
                        snippet = d.find("p", class_="post_intro_line")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('granddesignsmagazineScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class cabmScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.cottagesandbungalowsmag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post-masonry")

                    data = []
                    for d in data1:
                        title = d.find("h2", "entry-title")
                        snippet = d.find("div", class_="post-excerpt")
                        url = title.find("a")["href"]
                        short_description = d.find("ul", class_="post-categories")

                        short_description = short_description.text
                        short_description = wordninja.split(short_description)

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        sd = ""
                        for i in short_description:
                            sd = sd + " " + i

                        sd = str(sd).strip()
                        sd = ILLEGAL_CHARACTERS_RE.sub(r'', sd)

                        data.append({"short_description": sd, "title": title, "snippet": snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cabmScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class familyhandymanScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.familyhandyman.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="class-skills")

                    data = []
                    for d in data1:
                        title = d.find("h3", "entry-title")
                        snippet = d.find("div", class_="entry-summary")
                        url = title.find("a")["href"]
                        short_description = d.find("span", class_="cat-links")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append(
                            {"short_description": short_description, "title": title, "snippet": snippet,
                             "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('familyhandymanScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class tatlerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.tatler.com/search?q="+ seed.upper()
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="SummaryItemHedBase-dZZTtv fyYvOK summary-item__hed")
                        snippet = d.find("div", class_="BaseWrap-sc-UrHlS BaseText-fFrHpW SummaryItemDek-dwcsSh boMZdO eXFYcT dTrvxO summary-item__dek")
                        url = "https://www.tatler.com"+d.find("a", class_="SummaryItemHedLink-cgaOJy gsLgdZ summary-item-tracking__hed-link summary-item__hed-link summary-item__hed-link--underline-disable")["href"]
                        short_description = d.find("span", class_="RubricName-eXGqmo bHYiSS")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('tatlerScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class atomicranchScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.atomic-ranch.com/?s="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post-list")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        snippet = d.find("div", class_="post-excerpt")
                        url = title.find("a")["href"]
                        short_description = d.find("ul", class_="post-categories")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('atomicranchScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class nshoremagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.nshoremag.com/?s="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="cb-blog-style-a")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="cb-post-title")
                        snippet = d.find("div", class_="cb-excerpt")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="cb-taxonomy")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('nshoremagScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class homebuildingUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.homebuilding.co.uk/search?searchTerm="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")


                try:
                    data1 = soup.find_all("div", class_="listingResult")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="article-name")
                        snippet = d.find("p", class_="synopsis")
                        url = d.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('homebuildingUKScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class goodhousekeepingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.goodhousekeeping.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-q8qxmt epl65fo4")
                    print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-fq1qp3 e1rluvgc6")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ") +" & "+short_url.split("/")[2].replace("-", " ")
                        snippet= d.find("div", class_="css-ky7n27 e1rluvgc4")

                        url = "https://www.goodhousekeeping.com"+ short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('goodhousekeepingScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class goodhousekeepingUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.goodhousekeeping.com/uk/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item grid-simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title item-title")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[2].replace("-", " ")
                        snippet= d.find("div", class_="simple-item-dek item-dek")

                        url = "https://www.goodhousekeeping.com"+ short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('goodhousekeepingUKScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class minestomarketScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://minestomarket.news/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="tdb_module_loop")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        url = title.find("a")["href"]
                        short_description= d.find("span", class_="td-post-author-name")
                        snippet= d.find("div", class_="td-excerpt")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('minestomarketScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class newYorkerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.newyorker.com/search/q/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="River__riverItem___3huWr")

                    data = []
                    for d in data1:
                        links = d.find_all("a")
                        short_description = links[0]
                        title = d.find("h4", class_="River__hed___re6RP")
                        snippet = d.find("h5", class_="River__dek___CayIg")
                        url = "https://www.newyorker.com" + links[-2]["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('newYorkerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class moneySenseScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.moneysense.ca/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="post--item")

                    data = []
                    for d in data1:

                        short_description = d.find("p", class_="post-categories uppercase")
                        title = d.find("h3")
                        snippet = d.find("p", class_="excerpt")
                        url = d.find("a", class_="post--link")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}


                    print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('moneySenseScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class harpersBazaarScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.harpersbazaar.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-rrsgjg epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1ps51to e1rluvgc6")
                        snippet = d.find("div", class_="css-6tccnu e1rluvgc4")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ") + " & " + short_url.split("/")[
                            2].replace("-", " ")
                        url = "https://www.harpersbazaar.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('harpersBazaarScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class glamourScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.glamourmagazine.co.uk/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h3", class_="SummaryItemHedBase-dZZTtv eRdImi summary-item__hed")
                        snippet = ""
                        url = "https://www.glamourmagazine.co.uk" + d.find("a", class_="summary-item__hed-link")[
                            "href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('glamourScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class hgtvScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.hgtv.com/search/" + seed +"-"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("section", class_="o-ArticleResult o-ResultCard")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="m-MediaBlock__a-HeadlineText")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[3].replace("-", " ")
                        snippet= d.find("div", class_="m-MediaBlock__a-Description")
                        url= "https:"+short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('hgtvScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class houseBeautifulUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.housebeautiful.com/uk/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item")

                    data = []
                    for d in data1:
                        title = d.find("div", "simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[2].replace("-", " ")
                        url = "https://www.housebeautiful.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description":short_description, "title": title, "snippet": snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('houseBeautifulUKScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class tatlerasiaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.tatlerasia.com/search?q="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="card__item")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="font-style--heading")
                        url = "https://www.tatlerasia.com"+ title.find("a")["href"]
                        short_description = d.find("div", class_="category__item")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('tatlerasiaScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class homesandantiquesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.homesandantiques.com/search/?q="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="col-12 col-sm-6 col-md-4 template-search-universal__card")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="standard-card-new__display-title")
                        url = "https://www.homesandantiques.com"+ title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('homesandantiquesScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class luxurypoolsScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://luxurypools.com/?s="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="g1-collection-item")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="g1-gamma g1-gamma-1st entry-title")
                        snippet = d.find("div", class_="entry-summary")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('luxurypoolsScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class homesandinteriorsscotlandScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://homesandinteriorsscotland.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16 td_module_wrap td-animation-stack")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description,"title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('homesandinteriorsscotlandScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class californiahomedesignScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.californiahomedesign.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="col-sm-12 d-flex mb-4 border-bottom article post_content")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title search-post-title")
                        snippet = d.find("div", class_="entry-content")
                        url = title.find("a")["href"]
                        short_description =""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description,"title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('californiahomedesignScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class greenlivingmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://greenlivingmag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="tdb_module_loop td_module_wrap td-animation-stack td-cpt-post")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="td-module-meta-info")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('greenlivingmagScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class smartphotographyScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://smartphotography.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="entry")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        url = d.find("a")["href"]
                        snippet = d.find("div", class_="entry-content")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet = snippet.replace(" … ", "")
                        snippet = snippet.replace("[Read more...]", "")
                        snippet = snippet.replace("… Continue Reading", "")


                        data.append({"short_description": short_description,"title": title,"snippet": snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('smartphotographyScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class exhibitechScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.exhibit.tech/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1_temp = soup.find("div", class_="blog-posts-list blog-layout-Array")
                    data1 = data1_temp.find_all("article")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="post-title")
                        short_description = d.find("div", class_="post-categories")
                        url = title.find("a")["href"]
                        snippet=""

                        short_description = wordninja.split(short_description.text)
                        sd = ""
                        for i in short_description:
                            sd = sd + " " + i

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        sd = str(sd).strip()
                        sd = ILLEGAL_CHARACTERS_RE.sub(r'', sd)
                        sd = sd.replace("\n", "")
                        sd = sd.replace("\t", "")
                        sd = sd.replace("...", "")
                        sd = sd.replace("-", "")


                        data.append({"title": title,"short_description": sd, "snippet":snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('exhibitechScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class businesstravellerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.businesstraveller.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="latest_grid_item")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="top_related_title latest_item_title green_hover_target lat-itm-title")
                        short_description= d.find("div", class_="stripes")
                        url = d.find("a")["href"]
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description":short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('businesstravellerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class visiScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://visi.co.za/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="hoveraction")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        short_description= d.find("span", class_="entry-category")
                        url_tag= d.find("div", class_="entry-summary")
                        url = url_tag.find("a")["href"]
                        snippet = d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description": short_description, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('visiScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class gardenersmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://gardenersmag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="inside-article")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        url_tag= d.find("span", class_="cat-links")
                        url = title.find("a")["href"]
                        short_description = url_tag.find("a")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"title": title, "short_description": short_description, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gardenersmagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class sciencefocusScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.sciencefocus.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="col-12 col-sm-6 col-md-4 template-search-universal__card")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="standard-card-new__display-row")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet= short_url.split("/")[2].replace("-", " ")
                        url = "https://www.sciencefocus.com"+short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title,"short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('sciencefocusScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class whathifiScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.whathifi.com/search?searchTerm=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="listingResult")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="article-name")
                        url = d.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet= d.find("p", class_="synopsis")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title,"short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('whathifiScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class t3Scraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.t3.com/search?searchTerm=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="listing__link")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="listing__title")
                        url = d["href"]
                        short_description= url.split("/")[3].replace("-", " ")
                        snippet= d.find("p", class_="listing__text listing__text--strapline")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title,"short_description":short_description, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('t3Scraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class popularmechanicsScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.popularmechanics.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-l1en4j epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1eotuxi e1rluvgc6")
                        short_url = d["href"]
                        url="https://www.popularmechanics.com"+short_url
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet= d.find("div", class_="css-ghp3zv e1rluvgc4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title,"short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('popularmechanicsScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class wiredScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.wired.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo ejoBXF summary-item summary-item--has-border summary-item--article summary-item--no-icon summary-item--text-align-left summary-item--layout-placement-side-by-side-desktop-only summary-item--layout-position-image-left summary-item--layout-proportions-33-66 summary-item--side-by-side-align-top summary-item--side-by-side-image-right-mobile-true summary-item--standard SummaryItemWrapper-bGtGFH klkoMz summary-list__item")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="SummaryItemHedBase-dZZTtv fCrIUA summary-item__hed")
                        short_url = d.find("a", class_="SummaryItemHedLink-cgaOJy hYdAev summary-item-tracking__hed-link summary-item__hed-link")["href"]
                        url="https://www.wired.com"+short_url
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"title": title,"short_description":short_description , "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('wiredScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class pcmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.pcmag.com/search/results?query=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="w-full flex flex-wrap md:flex-nowrap py-4 border-b border-gray-lighter")

                    data = []
                    for d in data1:
                        title = d.find("h2")
                        short_url = title.find("a")["href"]
                        url= "https://www.pcmag.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet= d.find("p", class_="line-clamp-2 text-gray-darker")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title,"short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('pcmagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class retropopmagazineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://retropopmagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="elementor-post")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="elementor-post__title")
                        url = d.find("a")["href"]
                        short_description=""
                        snippet= d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('retropopmagazineScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class thetouriosityScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.thetouriosity.com/search-results?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="srskCNw oq5Rf88--fixed")

                    data = []
                    for d in data1:
                        title = d.find("a", class_="sil_d4M")
                        url = d.find("a")["href"]
                        short_description=""
                        snippet= d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('thetouriosityScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class traveltradejournalScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://traveltradejournal.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td-block-span6")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        url = title.find("a")["href"]
                        short_description=""
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)


                        data.append({"title": title,"short_description":short_description,"snippet":snippet ,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('traveltradejournalScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class gomagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "http://gomag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post-excerpt post")

                    data = []
                    for d in data1:
                        short_description= d.find("div", class_="entry-category")
                        title = d.find("h3", class_="entry-title")
                        snippet=""
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"title": title, "short_description": short_description, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gomagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class architecturalDigestScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.architecturaldigest.in/search/?q="+ seed.upper() +"&sort=score+desc"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo eOOpfj summary-item summary-item--has-border summary-item--article summary-item--no-icon summary-item--text-align-left summary-item--layout-placement-side-by-side-desktop-only summary-item--layout-position-image-left summary-item--layout-proportions-50-50 summary-item--side-by-side-align-center summary-item--side-by-side-image-right-mobile-false summary-item--standard SummaryItemWrapper-bGtGFH klkoMz summary-list__item")  #-------------

                    data = []
                    for d in data1:
                        short_description = d.find("span", class_="RubricName-eXGqmo bHYiSS")
                        title = d.find("h2", class_="SummaryItemHedBase-dZZTtv fqTtVh summary-item__hed")
                        snippet = d.find("div", class_="BaseWrap-sc-UrHlS BaseText-fFrHpW SummaryItemDek-dwcsSh boMZdO dHFxXE eWmA-dv summary-item__dek")
                        url = "https://www.architecturaldigest.in" + d.find("a", class_= "SummaryItemHedLink-cgaOJy jfeEmx summary-item-tracking__hed-link summary-item__hed-link")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('architecturalDigestScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class travelandleisureScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.travelandleisure.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="card__title")
                        url = d["href"]
                        short_description=url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"title": title, "short_description": short_description, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('travelandleisureScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class countryfileScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.countryfile.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="col-12 col-sm-6 col-md-4 template-search-universal__card")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="heading-4")
                        short_url = title.find("a")["href"]
                        url = "https://www.countryfile.com"+short_url
                        short_description=short_url.split("/")[1].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"title": title, "short_description": short_description, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('countryfileScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class indiaoutboundScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://indiaoutbound.info/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="gdlr-item gdlr-blog-grid")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="gdlr-blog-title")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet= d.find("div", class_="gdlr-blog-content ss")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")


                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description ,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('indiaoutboundScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class homemagazineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://homemagazine.nz/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="elementor-post")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="elementor-post__title")
                        url = d.find("a")["href"]
                        snippet = d.find("div", class_="elementor-post__excerpt")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('homemagazineScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class sahomeownerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.sahomeowner.co.za/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="search-item")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="title")
                        url = title.find("a")["href"]
                        snippet = d.find("div", class_="regularcontent")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description ,"snippet": snippet,
                                     "url": url})
                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('sahomeownerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class homesandgardensScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.homesandgardens.com/search?searchTerm=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="listing__item listing__item--alternate")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="listing__title")
                        url = d.find("a",class_="listing__link")["href"]
                        short_description = url.split("/")[3].replace("-", " ") + " "+ url.split("/")[4].replace("-", " ")
                        snippet = d.find("p", class_="listing__text listing__text--synopsis")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('homesandgardensScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class roadandtrackScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.roadandtrack.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1nzzxu9 epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1cu1gye e1rluvgc6")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.roadandtrack.com" + short_url
                        snippet = d.find("div", class_="css-atw22b e1rluvgc4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('roadandtrackScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class caranddriverScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.caranddriver.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1hqrmlz epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1vvw6c9 e1rluvgc6")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.caranddriver.com" + short_url
                        snippet = d.find("div", class_="css-1rp870h e1rluvgc4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('caranddriverScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class motortrendScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.motortrend.com/s/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="_3mDnE")

                    data = []
                    for d in data1:
                        title = d.find("a", class_="viYBm")
                        short_url = d.find("a", class_="viYBm")["href"]
                        short_description = d.find("div", class_="wQJal")
                        url = "https://www.motortrend.com" + short_url
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('motortrendScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class topgearmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.topgearmag.in/search?search=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="kjwrap tg-article-item")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="tg-review-title")
                        url_inside=d.find("div", class_="tg-media")
                        url = url_inside.find("a")["href"]
                        short_description = url.split("/")[4].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('topgearmagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class BCwomenUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://blackcountrywomensaid.co.uk/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)
                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="search-item")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        url = d.find("a")["href"]
                        snippet= d.find("p")
                        short_description=""
                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description": short_description,"snippet": snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('BCwomenUKScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class inkscapeScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://inkscape.org/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="result news")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet= d.find("p")
                        url = "https://inkscape.org"+ short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('inkscapeScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class bhgAUScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.bhg.com.au/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="ContentPreview")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="ContentPreview-Heading")
                        short_url = d.find("a")["href"]
                        short_description = d.find("a", class_="ContentPreview-Category")
                        snippet= d.find("p", class_="ContentPreview-Sell")
                        url = "https://www.bhg.com.au"+ short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                        #print(data2)

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('bhgAUScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class electronicsforuScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.electronicsforu.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td-block-span6")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        sd_url = d.find("div", class_="td-module-image")
                        sd_link = sd_url.find("a")["href"]
                        short_description = sd_link.split("/")[4].replace("-", " ")
                        url = title.find("a")["href"]
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"title": title,"short_description": short_description, "snippet":snippet ,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('electronicsforuScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
        wb.close()
        return True

class flowsmartlivingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://flowsmartliving.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="wgl_col-12 item")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="blog-post_title")
                        snippet = d.find("p")
                        short_description=""
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.replace("/", " ")
                        snippet = snippet.replace("+", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)

                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('flowsmartlivingScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class indiantextileScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.indiantextilemagazine.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("div", class_="col-sm-6 col-xxl-4 post-col")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        url = title.find("a")["href"]
                        short_description= d.find("div", class_="cat-links")
                        snippet= d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('indiantextileScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class houseandgardenUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.houseandgarden.co.uk/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-list__item")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="summary-item__hed")
                        short_url = d.find("a", class_="summary-item__hed-link--underline-disable")["href"]
                        short_description = d.find("a", class_="rubric__link")
                        snippet= d.find("div", class_="summary-item__dek")
                        url = "https://www.houseandgarden.co.uk"+ short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("_", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    print("..................")
                    print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('houseandgardenUKScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class urbanmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.urbanmag-online.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title")
                        url = title.find("a")["href"]
                        short_description= d.find("span", class_="cat-links")
                        snippet= d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('urbanmagScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class apparelScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://apparelmagazine.co.nz/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("article", class_="clearfix")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        url = title.find("a")["href"]
                        short_description= ""
                        snippet= d.find("p")
                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description":short_description ,"title": title, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('apparelScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class mensfolioScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.mens-folio.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("div", class_="card mb-4")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h5", class_="card-title")
                        url = d.find("a")["href"]
                        short_description= d.find("span", class_="date-posted date-posted-desktop")
                        snippet= d.find("p", class_="card-text")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({ "title": title, "short_description" : short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('mensfolioScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class femalemagSGScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.femalemag.com.sg/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("div", class_="story")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("a", class_="title")
                        url = d.find("a", class_="title")["href"]
                        short_description= d.find("a", class_="cat")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({ "title": title, "short_description" : short_description,"snippet":snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('femalemagSGScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class hairmagazineUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.hairmagazine.co.uk/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("li", class_="grid-style grid-overlay-meta")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="penci-entry-title entry-title grid-title")
                        url = title.find("a")["href"]
                        snippet= d.find("div", class_="item-content entry-content")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({ "title": title, "short_description":short_description,"snippet" : snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('hairmagazineUKScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class lofficielScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://lofficiel.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("div", class_="et_pb_column et_pb_column_1_2 free-archive-blog-individual free-archive-blog-normal")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="entry-title")
                        url = title.find("a")["href"]
                        snippet = d.find("div", class_="post-content")
                        short_description= d.find("span", class_="free-post-meta-category-extra free-background-category-main-color free-text-category-secondary-color")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = short_description.replace("(t)", "")
                        short_description = short_description.replace("(o)", "")
                        short_description = short_description.replace("(a)", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet = snippet.replace("READ MORE","")

                        data.append({ "title": title, "short_description" : short_description, "snippet":snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('lofficielScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class elleScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://elle.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)

                try:
                    data1 = soup.find_all("div", class_="elementor-column elementor-col-100 elementor-top-column elementor-element elementor-element-45f3f28d")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="elementor-heading-title elementor-size-default")
                        url = d.find("a")["href"]
                        short_description=""
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        data.append({ "title": title, "short_description":short_description, "snippet":snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('elleScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class cosmopolitanUKScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.cosmopolitan.com/uk/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="simple-item grid-simple-item grid-simple-item-last-mobile")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="simple-item-title")
                        snippet = d.find("div", class_="simple-item-dek")
                        short_url = d.find("a")["href"]
                        short_description = short_url.split("/")[2].replace("-", " ")
                        url = "https://www.cosmopolitan.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cosmopolitanUKScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class hotelbusinesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed
                #print(seed2)

                seed = seed.replace(" ", "+")

                URL = "https://hotelbusiness.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")
                #print(soup)


                try:
                    data1 = soup.find_all("div", class_="col-lg-6 mb-4")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title")
                        snippet = d.find("p")
                        short_description = d.find("div", class_="post-category text-uppercase py-2 px-3")
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('hotelbusinesScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class smarthomeworldScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.smarthomeworld.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="tdb_module_loop td_module_wrap td-animation-stack")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title td-module-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('smarthomeworldScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class psychologytodayScraper1(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.psychologytoday.com/us/archive?search=" + seed + "&undefined=Search+submit"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    ele=soup.find("ul", class_="grid--container--bb1 grid--container list-style-none")
                    data1 = ele.find_all("li")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h2")
                        url = d.find("a")["href"]
                        url= "https://www.psychologytoday.com"+url
                        snippet= d.find("div", class_="excerpt")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                        #print(data2)

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('psychologytodayScraper1.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class theEnglishGardenScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.theenglishgarden.co.uk/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="card mb-3 border-0 py-3")
                    #print(data1)

                    data = []
                    for d in data1:
                        title = d.find("h5", class_="card-title fw-bolder")
                        snippet = d.find("p", class_="card-text text-decoration-none text-black")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theEnglishGardenScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class newbeautyScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.newbeauty.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post-card")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="post-card__title")
                        url = d.find("a")["href"]
                        short_description=d.find("div", class_="post-card__category")
                        snippet= d.find("div", class_="post-card__byline")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet=snippet. replace('" ', " ")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    print(data2)

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('newbeautyScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class allureScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2=seed

                seed = seed.replace(" ", "+")

                URL = "https://www.allure.com/search?q=" + seed + "&sort=score+desc"

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL,headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_= "summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h3")
                        short_url=d.find("a", class_="summary-item__hed-link")["href"]
                        url= "https://www.allure.com"+short_url
                        snippet = d.find("div", class_="summary-item__dek")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet, "url":url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('allureScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class bhgScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.bhg.com/search/?q=" + seed + "&quotequery=&mod=DNH_S"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('bhgScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class cntravellerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.cntraveller.in/search/?q=" + seed + "&sort=score+desc"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("span", class_="RubricName-eXGqmo bHYiSS")
                        title = d.find("h2")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.cntraveller.in" + d.find("a", class_="summary-item__hed-link")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cntravellerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class cosmopolitanScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.cosmopolitan.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-qagrfk epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1dr0agl e1rluvgc6")
                        snippet = d.find("div", class_="css-1dlvc3a e1rluvgc4")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.cosmopolitan.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('cosmopolitanScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class dwellScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.dwell.com/query/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="_3mbrXRthE9F9t6iCkqG6F-")

                    data = []
                    for d in data1:
                        title = d
                        short_description=""
                        snippet = d.parent.find("div", class_="_37azNHMqzIJRxpdnCtt8J5")
                        url = "https://www.dwell.com" + title["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    print(data2)
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('dwellScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class gardensIllustratedScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.gardensillustrated.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="template-search-universal__card")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="standard-card-new__display-title")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        snippet=""
                        url = "https://www.gardensillustrated.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = re.sub(' +', ' ', title)
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gardensIllustratedScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class graziaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.grazia.co.in/search/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="thumbnail")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        url = title.find("a")["href"]
                        snippet = url.split("/")[3].replace("-", " ")
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"title": title, "short_description":short_description,"snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('graziaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class inStyleScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.instyle.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('inStyleScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class livingetcScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.livingetc.com/search?searchTerm=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="listing__item")

                    data = []
                    for d in data1:
                        url = d.find("a", class_="listing__link")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        title = d.find("h2", class_="listing__title")
                        snippet = d.find("p", class_="listing__text")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('livingetcScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class maximScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.maxim.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card-title")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('maximScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class realHomesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.realhomes.com/search?searchTerm=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="listing__item")

                    data = []
                    for d in data1:
                        url = d.find("a", class_="listing__link")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        title = d.find("h2", class_="listing__title")
                        snippet = d.find("p", class_="listing__text--strapline")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('realHomesScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class realSimpleScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.realsimple.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('realSimpleScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class styleAtHomeScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.styleathome.com/search?search[query]=" + seed + "&search[sort]="
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="result")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="title")
                        short_url = title.find("a")["href"]
                        url = "https://www.styleathome.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ") + " + " + short_url.split("/")[
                            2].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('styleAtHomeScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class southernLivingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.southernliving.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="card__content")
                        url = d["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")
                        short_description = ''.join([i for i in short_description if not i.isdigit()])

                        data.append({"short_description": short_description, "title": title,"snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('southernLivingScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class solitaireScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.solitairemagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="entry-item")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="entry-meta")
                        title = d.find("h2", class_="entry-title")
                        snippet = d.find("section", class_="entry-content")
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('solitaireScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class teenVogueScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.teenvogue.com/search?q=" + seed + "&sort=score+desc"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="summary-item__content")

                    data = []
                    for d in data1:
                        short_description = d.find("a", class_="RubricLink-DDpgX kWyAIu rubric__link")
                        title = d.find("h2", class_="summary-item__hed")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.teenvogue.com" + d.find("a", class_="summary-item__hed-link")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                         "url": url})

                        data2 = {"keyword": seed2, "data": data}
                        #print(data2)

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)


                    for each in json_without_slash['data']:
                        l = []
                        l.append(json_without_slash['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)

                        ws.append(data)

                    wb.save('teenVogueScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass

        wb.close()
        return True

class theJewelryScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.thejewelrymagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="td-module-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="td-module-meta-info")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet=snippet.replace("\r\n"," ")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theJewelryScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class womensHealthScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.womenshealthmag.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1mb238 epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-1xxuk0p e1rluvgc6")
                        snippet = d.find("div", class_="css-tzkl09 e1rluvgc4")
                        short_url = d["href"]
                        url = "https://www.womenshealthmag.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('womensHealthScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class vanityFairScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.vanityfair.com/search?q="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="SummaryItemWrapper-gcQMOo")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="summary-item__rubric")
                        title = d.find("h2")
                        snippet = d.find("div", class_="summary-item__dek")
                        url = "https://www.vanityfair.com" + d.find("a", class_="summary-item__hed-link")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('vanityFairScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class travelPeacockScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.travelpeacockmagazine.com/?s="+ seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="post")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="post-title")
                        url = d.find("a")["href"]
                        snippet = d.find("div", class_="post-content")
                        short_description = d.find("div", class_="thb-post-top")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('travelPeacockScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class travelAndLeisureIndiaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.travelandleisureindia.in/?s=" + seed + "&type=&category=&sort="
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="card")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="categories")
                        url = d.find("a")["href"]
                        snippet = url.split("/")[5].replace("-", " ")
                        short_description = url.split("/")[3].replace("-", " ") + url.split("/")[4].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", " ")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('travelAndLeisureIndiaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class townAndCountryScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.townandcountrymag.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-y7e8li epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-3wabvb e1rluvgc6")
                        snippet = d.find("div", class_="css-zdjp77 e1rluvgc4")
                        short_url = d["href"]
                        url = "https://www.townandcountrymag.com" + short_url
                        short_description = short_url.split("/")[1].replace("-", " ")+ " + " + short_url.split("/")[2].replace("-",
                                                                                                                        " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('townAndCountryScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class professionalWomanScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://professionalwomanmag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="type-post")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="entry-title")
                        snippet = d.find("div", class_="entry-content").find_all("p")[-1]
                        url = snippet.find("a", class_="more-link")["href"]
                        short_description = d.find("div", class_="entry-meta")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet=snippet.rstrip(" →")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('professionalWomanScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class motoringScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://motoringworld.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="card")

                    data = []
                    for d in data1:
                        content =d.find("div", class_="content")
                        title = content.find("div", class_="title")
                        url = title.find("a")["href"]
                        short_description = content.find("div", class_="section")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('motoringScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class moneyScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://money.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="search-result")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="headline")
                        url = "https://money.com" + title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description":short_description,"title": title, "snippet":snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('moneyScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class indianJewelerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://indianjeweller.in/search-news/" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", "news-row")

                    data = []
                    for d in data1:
                        title = d.find("h4", class_="title-16")
                        snippet = d.find("div", class_="desc-text")
                        url = title.find("a")["href"]
                        short_description = d.find("div", class_="post__small-text-meta").find_all("span")[1]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('indianJewelerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class stampingtonScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://stampington.com/search.php?search_query_adv=" + seed + "#search-results-information"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="card card-gallery card-hover")

                    data = []
                    for d in data1:
                        title = d.find("a", class_="card-ellipsis").span
                        url = d.find("a", class_="card-ellipsis")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('stampingtonScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class elleDecor2Scraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://elledecor.in/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_= "coloum col-sm-4")

                    data = []
                    for d in data1:
                        title = d.find("h4")
                        url = title.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description,"title": title, "snippet":snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('elleDecor2Scraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class flowerScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://flowermag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find("div", class_="blog blog-style-3x blog-style-masonry").find_all("div", class_= "description")

                    data = []
                    if len(data1) != 0:
                        for d in data1:
                            title = d.find("div", class_="post-title")
                            url = title.find("a")["href"]
                            snippet = d.find("div", class_="post-text")
                            short_description=""

                            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                            title = str(title.text).strip()
                            title = title.lstrip("\n\n\t\t\t\t")
                            title = title.rstrip("\n\t\t\t\n")
                            title = title.replace("–", "")
                            title = title.replace("|", "")
                            title = title.replace("—", "")
                            title = title.replace("#", "")
                            title = title.replace("+", "")
                            title = title.replace("'s", "")
                            title = title.replace("’s", "")
                            title = title.replace("‘", "")
                            title = title.replace("’", "")
                            title = title.replace("\n", "")
                            title = title.replace("\t", "")
                            title = title.replace("...", " ")
                            title = title.replace("…", "")
                            title = title.strip()
                            title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                            snippet = snippet.text.replace("\n", "")
                            snippet = snippet.replace("\t", "")
                            snippet = snippet.replace("\r", "")
                            snippet = snippet.replace("\xa0", " ")
                            snippet = snippet.strip()
                            snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                            snippet = snippet.replace("...", "")
                            snippet = snippet.replace("[…]", "")
                            snippet = snippet.replace("…", "")

                            data.append({"title": title, "snippet": snippet, "short_description":short_description,
                                         "url": url})

                            data2 = {"keyword": seed2, "data": data}
                        jsonData = json.dumps(data2)
                        json_without_slash = json.loads(jsonData)

                        s1 = json.dumps(json_without_slash)

                        json_object = json.loads(s1)
                        for each in json_object['data']:
                            l = []
                            l.append(json_object['keyword'])
                            l.append(each['title'])
                            l.append(each['short_description'])
                            l.append(each['snippet'])
                            l.append(each['url'])

                            x = each['title'].replace("...", "")
                            title_noun = TextBlob(x)
                            title_noun = title_noun.noun_phrases
                            l.append(str(title_noun))

                            data = tuple(l)
                            ws.append(data)

                        wb.save('flowerScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class gjepcScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://gjepc.org/solitaire/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find("div", class_="small-12 medium-8 columns").find_all("article", class_= "post")

                    data = []

                    if len(data1)!=0:
                        for d in data1:
                            short_description = d.find("aside", class_="post-meta")
                            title = d.find("div", class_="post-title")
                            url = title.find("a")["href"]
                            snippet = d.find("p")

                            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                            title = str(title.text).strip()
                            title = title.lstrip("\n\n\t\t\t\t")
                            title = title.rstrip("\n\t\t\t\n")
                            title = title.replace("–", "")
                            title = title.replace("|", "")
                            title = title.replace("—", "")
                            title = title.replace("#", "")
                            title = title.replace("+", "")
                            title = title.replace("'s", "")
                            title = title.replace("’s", "")
                            title = title.replace("‘", "")
                            title = title.replace("’", "")
                            title = title.replace("\n", "")
                            title = title.replace("\t", "")
                            title = title.replace("...", " ")
                            title = title.replace("…", "")
                            title = title.strip()
                            title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                            short_description = str(short_description.text).strip()
                            short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                            short_description = short_description.replace("\n", "")
                            short_description = short_description.replace("\t", "")
                            short_description = short_description.replace("...", "")
                            short_description = short_description.replace("-", "")

                            snippet = snippet.text.replace("\n", "")
                            snippet = snippet.replace("\t", "")
                            snippet = snippet.replace("\r", "")
                            snippet = snippet.replace("\xa0", " ")
                            snippet = snippet.strip()
                            snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                            snippet = snippet.replace("...", "")
                            snippet = snippet.replace("[…]", "")
                            snippet = snippet.replace("…", "")

                            data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                         "url": url})

                            data2 = {"keyword": seed2, "data": data}
                        jsonData = json.dumps(data2)
                        json_without_slash = json.loads(jsonData)

                        s1 = json.dumps(json_without_slash)

                        json_object = json.loads(s1)
                        for each in json_object['data']:
                            l = []
                            l.append(json_object['keyword'])
                            l.append(each['title'])
                            l.append(each['short_description'])
                            l.append(each['snippet'])
                            l.append(each['url'])

                            x = each['title'].replace("...", "")
                            title_noun = TextBlob(x)
                            title_noun = title_noun.noun_phrases
                            l.append(str(title_noun))

                            data = tuple(l)
                            ws.append(data)

                        wb.save('gjepcScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class instoreScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://instoremag.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="mvp-blog-story-out")

                    data = []
                    for d in data1:
                        link = d.find_all("a")[2]
                        short_description = d.find("div", class_="mvp-cat-date-wrap").find("a")
                        title = d.find("h2")
                        url = link["href"]
                        snippet = d.find("p")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("/", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('instoreScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class verandaScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.veranda.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1gij29w epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", class_="css-nukrc0 e1rluvgc6")
                        snippet =d.find("div",class_="css-tzkl09 e1rluvgc4")
                        short_url =d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.veranda.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('verandaScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class theSpruceScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")
                URL = "https://www.thespruce.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("li", class_="card-list__entry")

                    data = []
                    for d in data1:
                        short_description = d.find("div", class_="card__content").get('data-tag')
                        title = d.find("span", class_="card__title")
                        url = d.find("a", class_="card-list__card")["href"]
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('theSpruceScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class romanticHomesScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.romantichomes.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="post-list-bdr cf")

                    data = []
                    for d in data1:
                        title = d.find("div", class_="excerpt ex-wd").find("a")
                        url = title["href"]
                        short_description = url.split("/")[3].replace("-", " ") + " + " + url.split("/")[4].replace("-", " ")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('romanticHomesScraper.xlsx')


                except UnboundLocalError:

                    pass

                except AttributeError:

                    pass

                except KeyError:

                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class nylonScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.nylon.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="ofI")

                    data = []
                    for d in data1:
                        short_description = d.find("p", class_="pxx")
                        title = d.find("p", class_="icJ")
                        url = "https://www.nylon.com" + d["href"]
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('nylonScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class nationalGeographicScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.nationalgeographic.com/search?q=" + seed + "&location=srp&short_description=manual"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="ResultCard")

                    data = []
                    for d in data1:
                        title = d.find("span", class_="ResultCard__Title")
                        description = d.find("span", class_="ResultCard__Description")
                        snippet = description.find("span", class_="RichText")
                        url = d.find("a")["href"]
                        short_description = url.split("/")[3].replace("-", " ")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('nationalGeographicScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class lodgingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://lodgingmagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td-block-span6")

                    data = []
                    for d in data1:
                        short_description = d.find("a", class_="td-post-category")
                        title = d.find("h3")
                        snippet = d.find("div", class_="td-module-meta-info")
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('lodgingScraper.xlsx')

                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class houseBeautifulScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.housebeautiful.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-v7411v epl65fo4")

                    data = []
                    for d in data1:
                        title = d.find("h2", "css-1uklr3n e1rluvgc6")
                        snippet = d.find("div", class_="css-1lwfp25 e1rluvgc4")
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.housebeautiful.com" + short_url

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description":short_description, "title": title, "snippet": snippet, "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('houseBeautifulScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class hotelierScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.hoteliermagazine.com/?s=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("div", class_="td_module_16")

                    data = []
                    for d in data1:
                        title = d.find("h3", class_="entry-title")
                        snippet = d.find("div", class_="td-excerpt")
                        url = title.find("a")["href"]
                        short_description=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet=snippet.lstrip("\r\n                    \n")
                        snippet=snippet.rstrip("...                ")
                        snippet=snippet.replace("\n","")

                        data.append({"title": title, "short_description":short_description, "snippet": snippet,"url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('hotelierScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class gfmagScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.gfmag.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("article", class_="article-preview search-result")

                    data = []
                    for d in data1:
                        title = d.find("h3")
                        short_url = title.find("a")["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.gfmag.com" + short_url
                        snippet = d.find_all("p")[1]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")
                        snippet=snippet.replace("\r\n\r\n"," ")

                        short_description = short_description.strip("\n")

                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('gfmagScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class foodAndWineScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL = "https://www.foodandwine.com/search?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="comp mntl-card-list-items mntl-document-card mntl-card card card--no-image")

                    data = []
                    for d in data1:
                        url=d['href']
                        short_description = d.find("div", class_="card__content")["data-tag"]
                        title = d.find("span", class_="card__title-text")
                        snippet=""

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        data.append({"short_description": short_description, "title": title, "snippet":snippet,"url":url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('foodAndWineScraper.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class femina(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:

                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL="https://www.femina.in/search/tag_"+ seed +"&sort=score+desc?pg=10"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1= soup.find_all("div", class_="search-section")

                    data=[]
                    for d in data1:
                        short_description = d.find("div", class_="clearfix")
                        title = d.find("h3")
                        snippet = d.find("p")
                        url = title.find("a")["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")

                        data.append({"short_description":short_description,  "title":title ,"snippet": snippet, "url":url})

                        data2={"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('femina.xlsx')
                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

        wb.close()
        return True

class elleDecorScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")
                URL = "https://www.elledecor.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1xv7hpt epl65fo4")

                    data = []
                    for d in data1:
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.elledecor.com" + short_url
                        title = d.find("span", class_="css-1cjy3tu e1rluvgc5")
                        snippet = d.find("div", class_="css-7dr81s e1rluvgc4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")



                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('elleDecorScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class countryLivingScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2 = seed

                seed = seed.replace(" ", "+")

                URL ="https://www.countryliving.com/search/?q=" + seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1 = soup.find_all("a", class_="enk2x9t2 css-1bgh9hp epl65fo4")

                    data = []
                    for d in data1:
                        short_url = d["href"]
                        short_description = short_url.split("/")[1].replace("-", " ")
                        url = "https://www.countryliving.com" + short_url
                        title = d.find("span", class_="css-g0owdm e1rluvgc5")
                        snippet = d.find("div", class_="css-124muyb e1rluvgc4")

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet": snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}

                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    s1 = json.dumps(json_without_slash)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)

                    wb.save('countryLivingScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass
                except openpyxl.IllegalCharacterError:
                    pass

            wb.close()
        return True

class anOtherScraper(Resource):
    def get(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'word'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'short_description'
        ws.cell(row=1, column=4).value = 'snippet'
        ws.cell(row=1, column=5).value = 'url'
        ws.cell(row=1, column=6).value = 'title_noun_phrase'

        with open('input_file.txt', 'r') as file:
            for seed in file:
                seed = seed.partition(' :')
                seed = seed[0]
                seed = seed.strip("\n")
                seed = seed.strip()

                seed2=seed
                seed = seed.replace(" ", "+")

                URL="https://www.anothermag.com/search?q="+seed
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'}

                page = requests.get(URL, headers=headers, verify=False)

                soup = BeautifulSoup(page.content, "html.parser")

                try:
                    data1= soup.find_all("div",class_="article-list-item list-item")

                    data = []
                    for d in data1:
                        short_description = d.find("span", class_="section")
                        title = d.find("a", class_="title")
                        snippet=d.find("span",class_="nowrap")
                        url = "https://www.anothermag.com" + title["href"]

                        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                        title = str(title.text).strip()
                        title = title.lstrip("\n\n\t\t\t\t")
                        title = title.rstrip("\n\t\t\t\n")
                        title = title.replace("–", "")
                        title = title.replace("|", "")
                        title = title.replace("—", "")
                        title = title.replace("#", "")
                        title = title.replace("+", "")
                        title = title.replace("'s", "")
                        title = title.replace("’s", "")
                        title = title.replace("‘", "")
                        title = title.replace("’", "")
                        title = title.replace("\n", "")
                        title = title.replace("\t", "")
                        title = title.replace("...", " ")
                        title = title.replace("…", "")
                        title = title.strip()
                        title = ILLEGAL_CHARACTERS_RE.sub(r'', title)

                        short_description = str(short_description.text).strip()
                        short_description = ILLEGAL_CHARACTERS_RE.sub(r'', short_description)
                        short_description = short_description.replace("\n", "")
                        short_description = short_description.replace("\t", "")
                        short_description = short_description.replace("...", "")
                        short_description = short_description.replace("-", "")

                        snippet = snippet.text.replace("\n", "")
                        snippet = snippet.replace("\t", "")
                        snippet = snippet.replace("\r", "")
                        snippet = snippet.replace("\xa0", " ")
                        snippet = snippet.strip()
                        snippet = ILLEGAL_CHARACTERS_RE.sub(r'', snippet)
                        snippet = snippet.replace("...", "")
                        snippet = snippet.replace("[…]", "")
                        snippet = snippet.replace("…", "")


                        data.append({"short_description": short_description, "title": title, "snippet":snippet,
                                     "url": url})

                        data2 = {"keyword": seed2, "data": data}
                    jsonData = json.dumps(data2)
                    json_without_slash = json.loads(jsonData)

                    url = json_without_slash
                    s1 = json.dumps(json_without_slash)
                    # d2 = json.loads(s1)

                    json_object = json.loads(s1)
                    for each in json_object['data']:
                        l = []
                        l.append(json_object['keyword'])
                        l.append(each['title'])
                        l.append(each['short_description'])
                        l.append(each['snippet'])
                        l.append(each['url'])

                        x = each['title'].replace("...", "")
                        title_noun = TextBlob(x)
                        title_noun = title_noun.noun_phrases
                        l.append(str(title_noun))

                        data = tuple(l)
                        ws.append(data)
                    wb.save('anOtherScraper.xlsx')


                except UnboundLocalError:
                    pass
                except AttributeError:
                    pass
                except KeyError:
                    pass
                except TypeError:
                    pass
                except IndexError:
                    pass

        wb.close()
        return True

