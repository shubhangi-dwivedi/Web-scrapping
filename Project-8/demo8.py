import wordninja
from openpyxl.reader.excel import load_workbook
from bs4 import BeautifulSoup
import operator
import xlsxwriter
import xlwt
import xlrd
from xlutils.copy import copy

file1 = load_workbook(filename='job_record_python.xlsx')
ws=file1['python_jobs']

file2 = open("job_description.txt", "a")


start_col = 4
end_col = 4

for i in range(2, ws.max_row+1):
    row = [cell.value for cell in ws[i][start_col:end_col+1]]
    #print(row)

    string1 = ' '.join(row)
    #print(string1)

    sub1 = "JobDescription:_x000D_"
    sub2='...'
    #sub2 = "MoreDetails"
    idx1 = string1.index(sub1)
    idx2 = string1.index(sub2)

    res = ''
    # getting elements in between
    for idx in range(idx1 + len(sub1) + 1, idx2):
        res = res + string1[idx]

    #print(res)

    ans= wordninja.split(res)
    #print(ans)

    final_string=""
    for i in range(len(ans)):
        final_string= final_string+ans[i]+" "

    #print(final_string)
    file2.write(f'\n{final_string}')
    #file2.write("\n")

file2.close()
file1.close()