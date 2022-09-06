from openpyxl.reader.excel import load_workbook
import operator
import xlsxwriter

file1 = load_workbook(filename='job_records.xlsx')
ws=file1['python_jobs']

start_col = 2
end_col = 2

dict1={}

for i in range(2, ws.max_row+1):
    row = [cell.value for cell in ws[i][start_col:end_col+1]]
    string1=''

    for ele in row:
        string1+=ele

    res = string1.split(',')

    for i in res:

        # if there exists a key as "elements" then simply
        # increase its value.
        if i in dict1:
            dict1[i] += 1

        # if the dictionary does not have the key as "elements"
        # then create a key "elements" and assign its value to 1.
        else:
            dict1.update({i: 1})

dict1_sorted = dict(sorted(dict1.items(),key=operator.itemgetter(1),reverse=True))

workbook= xlsxwriter.Workbook("python_job_records"+".xlsx")
worksheet= workbook.add_worksheet("skills_records")
worksheet.write(0,0,'#')
worksheet.write(0,1,'skill')
worksheet.write(0,2,'freq')

i=0
for allKeys in dict1_sorted:
    worksheet.write(i+ 1, 0, str(i+ 1))
    worksheet.write(i + 1, 1, allKeys)
    worksheet.write(i + 1, 2, dict1[allKeys])
    i += 1

workbook.close()
