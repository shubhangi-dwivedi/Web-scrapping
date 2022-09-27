from openpyxl.reader.excel import load_workbook
import operator

file1 = load_workbook(filename='related_words1.xlsx')
ws=file1['synon']

start_col1 = 2
end_col1 = 4

dict1={}

a=0
for i in range(2, ws.max_row+1):
    a+=1
    print(a)
    row = [cell.value for cell in ws[i][start_col1:end_col1+1]]
    string1=''

    for ele in row:
        string1+=ele

    s1 = string1.replace("[", "")
    s1 = s1.replace("]", "")
    s1 = s1.replace("'", "")

    res = s1.split(',')

    for i in res:

        if i in dict1:
            dict1[i] += 1
        else:
            dict1.update({i: 1})

dict1_sorted = dict(sorted(dict1.items(),key=operator.itemgetter(1),reverse=True))

file1 = open("keywords1.txt", "a",encoding="utf-8")
for allKeys in dict1_sorted:
    txt=allKeys+' : '+ str(dict1[allKeys])
    print(txt)

    file1.write(txt)
    file1.write("\n")
    print(txt)

file1.close()