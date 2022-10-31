from openpyxl.reader.excel import load_workbook
import operator
# importing the module
from json import JSONDecodeError

import spacy
import wikipedia
import xlsxwriter
from textblob import TextBlob
import string
import re

from wikipedia import PageError, DisambiguationError

nlp = spacy.load("en_core_web_sm")

file1 = load_workbook(filename='cricket_wiki.xlsx')
ws=file1['wiki_scrap']

start_col1 = 3
end_col1 = 3

workbook = xlsxwriter.Workbook("wiki1" + ".xlsx")
worksheet = workbook.add_worksheet("wiki_scrap")
worksheet.write(0, 0, '#')
worksheet.write(0, 1, 'title')
worksheet.write(0, 2, 'url')
worksheet.write(0, 3, 'url_text')
worksheet.write(0, 4, 'url_text_ner')
worksheet.write(0, 5, 'number_of_ner')

index=0

for i in range(2, ws.max_row+1):
    row = [cell.value for cell in ws[i][start_col1:end_col1+1]]

    print(row)
    print(i)

    for ele in row:
        try:
            page_object1 = wikipedia.page(title=ele, pageid=None, auto_suggest=True, redirect=True, preload=False)
        # url of the page
            pg_url1 = page_object1.url
            print(pg_url1)
        # pg title
            pg_title1 = page_object1.original_title
        # pg link
            pg_link1 = page_object1.links
            print(pg_link1)

            title1 = ele
            for x in pg_link1:
                y = x.replace(" ", "_")
                print(y)

                try:
                    page_object2 = wikipedia.page(title=y, pageid=None, auto_suggest=True, redirect=True, preload=False)

                    # url of the page
                    pg_url2 = page_object2.url
                    print(pg_url2)
                    # pg title
                    pg_title2 = page_object2.original_title

                    blob_object = TextBlob(x)
                    pos_res = blob_object.tags

                    l = []

                    res = re.sub(r'[.,"\?:!;]', '', x)
                    doc = nlp(res)
                    entities = []
                    labels = []

                    print(doc.ents)

                    for v in doc.ents:
                        entities.append(v)
                        labels.append(v.label_)

                    l = list(zip(entities, labels))
                    print(l)

                    index += 1
                    print("index is : ", index)

                    worksheet.write(index, 0, index)
                    worksheet.write(index, 1, title1)
                    worksheet.write(index, 2, pg_url2)
                    worksheet.write(index, 3, pg_title2)
                    worksheet.write(index, 4, str(l))
                    worksheet.write(index, 5, len(l))

                    print("________________________")

                except PageError:
                    print("PageError exception")
                except DisambiguationError:
                    print("DisambiguationError exception")
                except JSONDecodeError:
                    print("JSONDecodeError exception")
                except KeyError:
                    print("KeyError exception")
                except DisambiguationError:
                    print("DisambiguationError exception")
                except ValueError:
                    print("ValueError exception")

        except PageError:
            print("PageError exception")
        except DisambiguationError:
            print("DisambiguationError exception")
        except JSONDecodeError:
            print("JSONDecodeError exception")
        except KeyError:
            print("KeyError exception")
        except DisambiguationError:
            print("DisambiguationError exception")
        except ValueError:
            print("ValueError exception")



workbook.close()
